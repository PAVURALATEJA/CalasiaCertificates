"""
Generate a sample Excel template for instrument bulk import.
Run this once to create the template file: instruments_template.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Instruments"

# Header styling
header_fill   = PatternFill("solid", fgColor="1A2340")
required_fill = PatternFill("solid", fgColor="F97316")
optional_fill = PatternFill("solid", fgColor="22D3EE")
white_font    = Font(bold=True, color="FFFFFF", size=11)
thin_border   = Border(
    left=Side(style='thin', color='1E2D50'),
    right=Side(style='thin', color='1E2D50'),
    top=Side(style='thin', color='1E2D50'),
    bottom=Side(style='thin', color='1E2D50')
)

headers = [
    ("company_name",     "REQUIRED", 22),
    ("instrument_name",  "REQUIRED", 24),
    ("serial_number",    "REQUIRED", 20),
    ("model",            "optional", 18),
    ("manufacturer",     "optional", 20),
    ("location",         "optional", 20),
]

for col_idx, (col_name, req, width) in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.font = white_font
    cell.fill = required_fill if req == "REQUIRED" else PatternFill("solid", fgColor="0891B2")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    ws.column_dimensions[cell.column_letter].width = width

ws.row_dimensions[1].height = 28

# Sample data rows
sample_data = [
    ("Qualcomm India",   "Oscilloscope",          "B012036",  "TAP3500",   "Keysight",   "Hyderabad Lab"),
    ("Qualcomm India",   "Communication Analyzer", "165067",   "CMW500",    "Rohde&Schwarz", "Hyderabad Lab"),
    ("Acme Corp",        "Multimeter",             "DMM-0012", "34465A",    "Keysight",   "Pune Office"),
    ("Acme Corp",        "Power Supply",           "PS-9901",  "E3631A",    "Keysight",   "Pune Office"),
    ("Global Tech Ltd",  "Signal Generator",       "SG-445",   "N5181B",    "Keysight",   "Delhi Site"),
    ("Global Tech Ltd",  "Spectrum Analyzer",      "SA-112",   "N9020B",    "Keysight",   "Delhi Site"),
    ("Infosys Systems",  "LCR Meter",              "LCR-221",  "E4980AL",   "Keysight",   "Bengaluru"),
    ("Infosys Systems",  "Function Generator",     "FG-8844",  "33500B",    "Keysight",   "Bengaluru"),
]

row_colors = ["0F1629", "141D35"]
for row_idx, row in enumerate(sample_data, start=2):
    fill = PatternFill("solid", fgColor=row_colors[row_idx % 2])
    for col_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = fill
        cell.font = Font(color="E2E8F0", size=10)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center')

# Add note row
note_row = ws.max_row + 2
ws.cell(row=note_row, column=1, value="📌 NOTE: Row 1 must be the header (as shown above). Orange columns are required. Add your data from row 2 onwards.")
ws.cell(row=note_row, column=1).font = Font(color="F59E0B", size=9, italic=True)
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=6)

out_path = r"c:\Users\Dell 5400\Desktop\calasiacertificates\instruments_template.xlsx"
wb.save(out_path)
print(f"Template saved to: {out_path}")
print("Upload this file at: http://127.0.0.1:5000/admin/instruments/import-excel")
